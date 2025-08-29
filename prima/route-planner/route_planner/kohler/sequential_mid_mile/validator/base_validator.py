from route_planner.utils import logging
import mimetypes
import re

import pandas as pd
from datetime import datetime
from pandas.api.types import is_string_dtype
from openpyxl import load_workbook

from route_planner.constants.app_constants import VALID_FILE_EXTENSIONS, BP_MID_MILE_SHEET_NAMES, NUMBER_DTYPE_KIND, VALID_UOM_TYPE_LIST
from route_planner.exceptions.exceptions import InvalidRequest, ValidationError
from route_planner.vrp.sku_fixed_cost_planner.validator import utils
from werkzeug.exceptions import BadRequest

logger = logging.getLogger(__name__)


class BaseValidator(object):

    def __init__(self, file=None, rid=None):
        logger.info(":: BaseValidator init::")
        # FileStorage Object
        # self.file = self.validate_file(file)
        self.file = file
        self.rid = rid
        self.validate_sheets()

    def validate_file(self, file):
        if not file:
            raise InvalidRequest(message="No file provided")

        mime_type = mimetypes.guess_type(file.filename)

        if not mime_type[0]:
            raise InvalidRequest(message="Invalid Request File")

        file_ext = mimetypes.guess_extension(mime_type[0])
        if file_ext not in VALID_FILE_EXTENSIONS:
            raise InvalidRequest(
                message=f"Invalid File Extension '{file_ext}', Valid Extensions: {VALID_FILE_EXTENSIONS}")

        return file

    def validate_sheets(self):
        # open an Excel file and return a workbook
        wb = load_workbook(self.file, read_only=True)
        if not set(BP_MID_MILE_SHEET_NAMES).issubset(set(wb.sheetnames)):
            raise BadRequest(f"Sheet not found!, Valid Sheets: {BP_MID_MILE_SHEET_NAMES}")

    @staticmethod
    def validate_header(validator) -> bool:
        """
        Validate validator Headers
        """
        valid_header = validator.VALID_HEADER.copy()
        actual_header = validator.header.copy()
        if set(valid_header) != set(actual_header):
            rows = utils.get_header_error_response(valid_header, actual_header)
            message = f"Invalid {validator.SHEET} Headers!"
            validator.problems.extend(BaseValidator.add_problem([-1], rows, message, validator.SHEET))
            return False
        return True

    @staticmethod
    def add_problem(idx, rows, message, sheet):
        problem = []
        if not rows and len(idx) == 1:
            problem.append(dict(row_number=idx[0] + 2, row=dict(), message=message, sheet=sheet))
            return problem

        x = pd.DataFrame(rows).fillna("").to_json(orient="records")
        rows = pd.read_json(x, orient="records").to_dict(orient='records')
        for i, row in zip(idx, rows):
            # 2 = 0-indexing + header-row
            problem.append(dict(row_number=i + 2, row=row, message=message, sheet=sheet))
        return problem

    @staticmethod
    def check_MandatoryField(self):
        valid = True
        for col in self.MANDATORY_COLS:
            check_null = self.df[col].isnull()
            if not check_null.all() and check_null.any():
                idx = self.df[check_null].index.to_list()
                rows = self.df.loc[idx].to_dict(orient="records")
                message = f"Mandatory field ({col}) not provided"
                self.problems.extend(BaseValidator.add_problem(idx, rows, message, self.SHEET))
                valid = False
            elif check_null.all():
                message = f"Mandatory Column ({col}) not provided"
                self.problems.extend(BaseValidator.add_problem([-1], list(), message, self.SHEET))
                valid = False

    @staticmethod
    def number_type_validator(self, columns):
        invalid_cols = []
        for col_name in columns:
            if self.df[col_name].dtype.kind not in NUMBER_DTYPE_KIND:
                invalid_cols.append(col_name)

        if invalid_cols:
            message = f"Number Type Validation Failed for {','.join(invalid_cols)}"
            self.problems.extend(BaseValidator.add_problem([-1], dict(), message, self.SHEET))
            raise ValidationError(problems=self.problems)

    # @staticmethod
    # def string_type_validator(self):
    #     invalid_cols = list()
    #     dtype_map = utils.true_dtype(self.df[self.STRING_TYPE_COLS])
    #
    #     for col_name, actual_types in dtype_map.items():
    #         valid_bool = [True if a_type == str else False for a_type in actual_types]
    #         if not all(valid_bool):
    #             invalid_cols.append(col_name)
    #
    #     if invalid_cols:
    #         message = f"String Type Validation Failed for ({','.join(invalid_cols)})"
    #         self.problems.extend(BaseValidator.add_problem([-1], dict(), message, self.SHEET))
    #         raise ValidationError(problems=self.problems)
    @staticmethod
    def string_type_validator(self, cols):
        invalid_cols = list()
        for col in cols:
            if not is_string_dtype(self.df[col]):
                invalid_cols.append(col)

        if invalid_cols:
            message = f"String Type Validation Failed for ({', '.join(invalid_cols)})"
            self.problems.extend(BaseValidator.add_problem([-1], dict(), message, self.SHEET))

    @staticmethod
    def datetime_type_validator(self):
        self.df["Placement Date & Time"] = self.df["Placement Date & Time"].fillna(datetime.now().strftime("%d-%m-%Y %H:%M"))
        cols1 = self.df[pd.to_datetime(self.df["Placement Date & Time"], format='%d-%m-%Y %H:%M', errors='coerce').isna()]
        indices = cols1.index.tolist()
        rows = self.df.loc[indices].to_dict(orient='records')
        if rows:
            message = f"Datetime Type Validation Failed for these rows"
            self.problems.extend(BaseValidator.add_problem(indices, rows, message, self.SHEET))
        else:
            current_time = datetime.now().strftime("%d-%m-%Y %H:%M")
            a = self.df[self.df["Placement Date & Time"] < current_time]
            indices1 = a.index.values.tolist()
            rows1 = self.df.loc[indices1].to_dict(orient='records')
            if rows1:
                message = f"Datetime Type Validation Failed for these rows - enter future datetime"
                self.problems.extend(BaseValidator.add_problem(indices1, rows1, message, self.SHEET))

    @staticmethod
    def validate_coordinate(self, col, type, drop_null=False):
        if drop_null:
            col = col.dropna()

        COORDINATES = {'latitude': {'min_range': -90, 'max_range': 90},
                       'longitude': {'min_range': -180, 'max_range': 180}
                       }
        coordinates_range = COORDINATES.get(type)
        min_range_bool = col >= coordinates_range.get('min_range')
        max_range_bool = col <= coordinates_range.get('max_range')

        if not (min_range_bool & max_range_bool).all():
            index_bool = (min_range_bool & max_range_bool)==False
            indexes = col[index_bool].index.tolist()
            message = f"{col.name} does not satisfy constraint range {coordinates_range['min_range'], coordinates_range['max_range']}"
            rows = self.df.loc[indexes].to_dict(orient="records")
            self.problems.extend(BaseValidator.add_problem(indexes, rows, message, self.SHEET))

    @staticmethod
    def positive_number_validator(self, columns):
        error = False
        for column in columns:
            col_not_null = self.df[column].notnull()
            col = self.df[col_not_null][column]
            if (col < 0).any():
                indexes = col[col < 0].index
                rows = self.df.loc[indexes].to_dict(orient='records')
                message = f"{column} value should be >= 0"
                self.problems.extend(BaseValidator.add_problem(indexes, rows, message, self.SHEET))
                error = True
        if error:
            raise ValidationError(problems=self.problems)

    @staticmethod
    def validate_uom(self):
        col = self.df['UOM* (kg, cbm)']
        check_col = col.copy().str.lower()
        error = False

        invalid_bool = check_col.isin(VALID_UOM_TYPE_LIST)
        if not invalid_bool.all():
            error = True
            indexes = invalid_bool == False
            rows = self.df[indexes].to_dict(orient='records')
            message = f"Invalid Value for UOM* (kg, cbm)"
            self.problems.extend(BaseValidator.add_problem(indexes, rows, message, self.SHEET))

        unique_bool = len(check_col.unique()) == 1
        if not unique_bool:
            error = True
            rows = list()
            rows.append(col.to_dict())
            message = f"UOM* (kg, cbm) should be unique through out"
            self.problems.extend(BaseValidator.add_problem([-1], rows, message, self.SHEET))
            # logger.info(rows)

        if not error:
            self.UOM_TYPE = col[0]

    @staticmethod
    def df_to_csv(self):
        self.new_df.to_csv(f"{self.rid}_{self.SHEET}.csv", index=False)

    @staticmethod
    def greater_than_zero_validation(validator, columns: list) -> bool:
        error = False
        for col_name in columns:
            col = validator.df[col_name].copy().dropna()
            invalid_bool = col == 0


            if invalid_bool.any():
                indexes = col[invalid_bool].index.to_list()
                rows = validator.df.loc[indexes].to_dict(orient='records')
                message = f"{col.name} cannot have 0 value"
                validator.problems.extend(BaseValidator.add_problem(indexes, rows, message, validator.SHEET))
                error = True

        return error

    @staticmethod
    def strip_whitespace(string):
        string = re.sub(r"\s+", "", string, flags=re.UNICODE)
        return string

    @staticmethod
    def validate_comma_sep_string_col(column: pd.Series) -> pd.Series:
        col = column.copy()
        col = col.fillna('')
        col = col.apply(lambda x: BaseValidator.strip_whitespace(x))
        return col

    @staticmethod
    def parse_case_insensitive_distinct(column: pd.Series) -> pd.Series:
        """
        for commma sep string
        """
        col = column.copy()
        col = col.map(lambda elements: ','.join(set((elements.lower()).split(','))))
        return col

    @staticmethod
    def check_priority_value(self, column):
        col_not_null = self.df[column].notnull()
        col = self.df[col_not_null][column]
        if (col > 100).any():
            indexes = col[col > 100].index
            rows = self.df.loc[indexes].to_dict(orient='records')
            message = f"{column} value should be <= 100"
            self.problems.extend(BaseValidator.add_problem(indexes, rows, message, self.SHEET))

    @staticmethod
    def type_cast_to_str(validator):
        invalid_cols = list()
        for col_name in validator.STRING_TYPE_COLS:
            col = validator.df[col_name].copy().dropna()
            col = col.apply(str)
            indexes = col.index
            validator.df[col_name].iloc[indexes] = col

    @staticmethod
    def string_conversion(col):
        return str(str(col).strip())

    @staticmethod
    def int_conversion(col, df):
        df[col] = df[col].astype(int)

    @staticmethod
    def check_white_spaces(self):
        for col in self.VALID_HEADER:
            s = pd.Series(self.df[col].apply(str))
            s = s.str.isspace()
            indexes = (s[s == 1].index.values).tolist()
            if len(indexes):
                rows = self.df.loc[indexes].to_dict(orient='records')
                message = f"Found spaces as values in {col}"
                self.problems.extend(BaseValidator.add_problem(indexes, rows, message, self.SHEET))

    @staticmethod
    def validate_alpha_count(validator, columns: list, min_alpha_count=4) -> bool:
        """
           Validates the minimum count of alphabetic characters in specified columns of a DataFrame.

           Args:
               validator: BaseValidator object
               columns (list): List of column names to be validated.
               min_alpha_count (int, optional): Minimum count of alphabetic characters. Default is 4.

           Returns:
               bool: Is valid or not
        """
        error = False
        pattern = r'[a-zA-Z]'
        for col_name in columns:
            col = validator.df[col_name].copy().fillna('')
            alpha_count = col.str.count(pattern, flags=re.IGNORECASE)
            invalid_bool = alpha_count < min_alpha_count

            if invalid_bool.any():
                indexes = col[invalid_bool].index.to_list()
                rows = validator.df.loc[indexes].to_dict(orient='records')
                message = f"{col.name} requires at-least {min_alpha_count} alphabets"
                validator.problems.extend(BaseValidator.add_problem(indexes, rows, message, validator.SHEET))
                error = True

        return error

    @staticmethod
    def validate_file_length(validator):
        """
        Validate file length
        """
        file_length_limit = validator.FILE_ROWS_LIMIT + 1

        if validator.df.shape[0] >= file_length_limit:
            message = f"The maximum number of {validator.SHEET} allowed is {validator.FILE_ROWS_LIMIT}"
            problem = BaseValidator.add_problem([-1], None, message, validator.SHEET)
            validator.problems.extend(problem)
            return False
        return True

