from route_planner.utils import logging
import re

import pandas as pd
from openpyxl import load_workbook

from route_planner.constants.app_constants import VALID_SHEET_NAMES, NUMBER_DTYPE_KIND, \
    VALID_UOM_TYPE_LIST
from route_planner.exceptions.exceptions import InvalidRequest, ValidationError
from route_planner.vrp.sku_fixed_cost_planner.validator import utils

logger = logging.getLogger(__name__)


class BaseValidator(object):

    def __init__(self, file=None, rid=None):
        logger.info(":: BaseValidator init::")
        self.file = file
        self.rid = rid
        self.validate_sheets()

    def validate_sheets(self):
        # open an Excel file and return a workbook
        wb = load_workbook(self.file, read_only=True)
        if not set(VALID_SHEET_NAMES).issubset(set(wb.sheetnames)):
            raise InvalidRequest(message=f"Sheet not found!, Valid Sheets: {VALID_SHEET_NAMES}")

    @staticmethod
    def validate_header(validator) -> bool:
        """
        Validate validator Headers
        """
        valid_header = validator.VALID_HEADER.copy()
        actual_header = validator.header.copy()
        if set(valid_header) != set(actual_header):
            rows = utils.get_header_error_response(valid_header, actual_header)
            message = f"Invalid {validator.SHEET} Headers!"
            validator.problems.extend(BaseValidator.add_problem([-1], rows, message, validator.SHEET))
            return False
        return True

    @staticmethod
    def add_problem(idx, rows, message, sheet):
        problem = []
        if not rows and len(idx) == 1:
            problem.append(dict(row_number=idx[0] + 2, row=dict(), message=message, sheet=sheet))
            return problem

        x = pd.DataFrame(rows).fillna("").to_json(orient="records")
        rows = pd.read_json(x, orient="records").to_dict(orient='records')
        for i, row in zip(idx, rows):
            # offset is 2 because 0-indexing + header-row
            problem.append(dict(row_number=i + 2, row=row, message=message, sheet=sheet))
        return problem

    @staticmethod
    def check_mandatory_fields(validator):
        for col in validator.MANDATORY_COLS:
            check_null = validator.df[col].isnull()
            if not check_null.all() and check_null.any():
                idx = validator.df[check_null].index.to_list()
                rows = validator.df.loc[idx].to_dict(orient="records")
                message = f"Mandatory field ({col}) not provided"
                validator.problems.extend(BaseValidator.add_problem(idx, rows, message, validator.SHEET))
            elif check_null.all():
                message = f"Mandatory Column ({col}) not provided"
                validator.problems.extend(BaseValidator.add_problem([-1], list(), message, validator.SHEET))

    @staticmethod
    def number_type_validator(validator):
        invalid_cols = []
        for col_name in validator.NUMBER_TYPE_COLS:
            if validator.df[col_name].dtype.kind not in NUMBER_DTYPE_KIND:
                invalid_cols.append(col_name)

        if invalid_cols:
            message = f"Number Type Validation Failed for {','.join(invalid_cols)}"
            validator.problems.extend(BaseValidator.add_problem([-1], dict(), message, validator.SHEET))
            raise ValidationError(problems=validator.problems)

    @staticmethod
    def string_type_validator(validator):
        invalid_cols = list()
        dtypes_map = utils.true_dtype(validator.df[validator.STRING_TYPE_COLS])

        for col_name, actual_types in dtypes_map.items():
            valid_bool = [True if a_type == str else False for a_type in actual_types]
            if not all(valid_bool):
                invalid_cols.append(col_name)

        if invalid_cols:
            message = f"String Type Validation Failed for ({','.join(invalid_cols)})"
            validator.problems.extend(BaseValidator.add_problem([-1], dict(), message, validator.SHEET))
            raise ValidationError(problems=validator.problems)

    @staticmethod
    def validate_coordinate(validator, _col, type, mandatory_col=True):
        _coordinates = {'latitude': {'min_range': -90, 'max_range': 90},
                        'longitude': {'min_range': -180, 'max_range': 180}
                        }
        col = _col.copy()
        if not mandatory_col:
            col = col.dropna()

        coordinates_range = _coordinates.get(type)
        min_range_bool = col >= coordinates_range.get('min_range')
        max_range_bool = col <= coordinates_range.get('max_range')

        if not (min_range_bool & max_range_bool).all():
            indexes = (min_range_bool & max_range_bool) == False
            message = f"{col.name} does not satisfy constraint range {coordinates_range['min_range'], coordinates_range['max_range']}"
            rows = validator.df[indexes].to_dict(orient="records")
            validator.problems.extend(BaseValidator.add_problem(indexes, rows, message, validator.SHEET))

    @staticmethod
    def positive_number_validator(validator, columns, greater_than_zero=False, fill_na=None):
        error = False
        for column in columns:
            if fill_na is not None:
                validator.df[column] = validator.df[column].fillna(fill_na)
            col_not_null = validator.df[column].notnull()
            col = validator.df[col_not_null][column]
            if greater_than_zero and (col <= 0).any():
                indexes = col[col <= 0].index
                rows = validator.df.loc[indexes].to_dict(orient='records')
                message = f"{column} value should be > 0"
                validator.problems.extend(BaseValidator.add_problem(indexes, rows, message, validator.SHEET))
                error = True
            elif (col < 0).any():
                indexes = col[col < 0].index
                rows = validator.df.loc[indexes].to_dict(orient='records')
                message = f"{column} value should be >= 0"
                validator.problems.extend(BaseValidator.add_problem(indexes, rows, message, validator.SHEET))
                error = True
        if error:
            raise ValidationError(problems=validator.problems)

    @staticmethod
    def validate_uom(validator):
        col = validator.df['UOM* (kg, cbm)']
        check_col = col.copy().str.lower()
        error = False

        invalid_bool = check_col.isin(VALID_UOM_TYPE_LIST)
        if not invalid_bool.all():
            error = True
            indexes = invalid_bool == False
            rows = validator.df[indexes].to_dict(orient='records')
            message = f"Invalid Value for UOM* (kg, cbm)"
            validator.problems.extend(BaseValidator.add_problem(indexes, rows, message, validator.SHEET))

        unique_bool = len(check_col.unique()) == 1
        if not unique_bool:
            error = True
            rows = list()
            rows.append(col.to_dict())
            message = f"UOM* (kg, cbm) should be unique through out"
            validator.problems.extend(BaseValidator.add_problem([-1], rows, message, validator.SHEET))

        if not error:
            validator.UOM_TYPE = col[0]

    @staticmethod
    def df_to_csv(validator):
        validator.new_df.to_csv(f"{validator.rid}_{validator.SHEET}.csv", index=False)

    @staticmethod
    def strip_whitespace(string):
        string = re.sub(r"\s+", "", string, flags=re.UNICODE)
        return string

    @staticmethod
    def validate_comma_sep_string_col(column: pd.Series) -> pd.Series:
        col = column.copy()
        col = col.apply(lambda x: BaseValidator.strip_whitespace(x)if type(x) == str else x)
        return col

    @staticmethod
    def parse_case_insensitive_distinct(column: pd.Series) -> pd.Series:
        """
        for comma sep string
        """
        col = column.copy()
        col = col.map(lambda e: ','.join(set((e.lower()).split(',')))if type(e) == str else e)
        return col

    @staticmethod
    def type_cast_to_str(df: pd.DataFrame, cols: list, fill=""):
        valid_cols = df.columns
        actual_cols = list(valid_cols.intersection(cols))
        if actual_cols and fill is not None:
            df[actual_cols] = df[actual_cols].fillna(fill).astype(str)
        elif actual_cols:
            df[actual_cols] = df[actual_cols].astype(str)

    @staticmethod
    def from_to_city_validation(validator, from_city, to_city):

        # value should exist in both To City and From City
        x = validator.df[from_city].copy().notnull()
        y = validator.df[to_city].copy().notnull()
        working_window_bool = x ^ y
        if working_window_bool.any():
            # for 01 and 10 case
            indexes = working_window_bool[working_window_bool == True].index.to_list()
            rows = validator.df.loc[indexes].to_dict(orient="records")
            message = f"Invalid both {from_city} and {to_city} should be provided!"
            validator.problems.extend(BaseValidator.add_problem(indexes, rows, message, validator.SHEET))
            return False

        validator.new_df['from_city'] = validator.df['From City'].fillna('').str.lower()
        validator.new_df['to_city'] = validator.df['To City'].fillna('').str.lower()
        return True
