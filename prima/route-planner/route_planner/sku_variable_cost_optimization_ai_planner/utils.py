import math
import uuid
from datetime import datetime, timedelta
from openpyxl.utils.cell import get_column_letter
import numpy as np
import pandas as pd
import pytz
from solvers.solver_v3.constants import UTC_FORMAT


def _get_local_timezone():
    return pytz.timezone('Asia/Kolkata')


def generate_request_id():
    return str(uuid.uuid4()).replace('-', '')


def get_current_timestamp():
    return datetime.now(tz=_get_local_timezone())


def dataframe_empty(df_column):
    empty_list = df_column.dropna().empty
    return empty_list


def get_now() -> datetime:
    return datetime.now(pytz.timezone('Asia/Calcutta'))


def get_eta(now: datetime, time_taken: int, time_format: str = '%d/%m/%Y %H:%M'):
    eta = now + timedelta(hours=time_taken)
    eta = eta.strftime(time_format)
    return eta


def type_cast_to_str(df: pd.DataFrame, str_type_cols: list) ->None:
    if df:
        df_columns = df.columns.to_list()
        for col_name in str_type_cols:
            if col_name in df_columns:
                col = df[col_name].copy().dropna()
                col = col.apply(str)
                indexes = col.index
                df[col_name].iloc[indexes] = col


def get_column_letter_by_index(value: int) -> str:
    """
    Return excel column index for given numeric zero based index.

    Example:
        0 -> A
    """
    # Todo refactor
    # __ALPHA = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    # def hash_num_to_char(number: int) -> str:
    #     """
    #     Recursive function to get excel header by index.
    #
    #     if number is less than "26", simply hash out (index-1)
    #     Remaining possibilities,
    #     1. if remainder is zero (if quotient is 1 or not 1)
    #     2. if remainder is not zero
    #     """
    #     if number < 26:
    #         return __ALPHA[number - 1]
    #
    #     q, r = number // 26, number % 26
    #
    #     if r == 0 and q == 1:
    #         return __ALPHA[r - 1]
    #     elif r == 0:
    #         return hash_num_to_char(q - 1) + __ALPHA[r - 1]
    #     else:
    #         return hash_num_to_char(q) + __ALPHA[r - 1]
    #
    # return hash_num_to_char(value + 1)

    return get_column_letter(value + 1)


def get_header_error_response(valid_header: list, actual_header: list) -> list:
    """
    Generate Header Error Response.

    {Column index : {'Actual Header' : actual_column_name, 'Valid Header' : valid_column_name}}
    Example:
        {
            "H": {
                "Actual Header": "Unnamed: 7",
                "Valid Header": "Loading Time (Hours)"
            },
            "O": {
                "Actual Header": "Extra",
                "Valid Header": "Serviceable Vehicles"
            },
            "P": {
                "Actual Header": "Extra2",
                "Valid Header": ""
            }
        }
    """

    response = list()
    size_valid = len(valid_header)
    size_invalid = len(actual_header)
    longer_list, shorter_list = (valid_header, actual_header) if size_valid > size_invalid else (actual_header, valid_header)

    size_delta = len(longer_list) - len(shorter_list)
    shorter_list.extend([""] * size_delta)

    response_dict = dict()
    for idx, (valid_value, actual_value) in enumerate(zip(valid_header, actual_header)):
        temp_dict = dict()
        temp_dict["Actual Header"] = actual_value
        temp_dict["Valid Header"] = valid_value
        index = get_column_letter_by_index(idx)

        if actual_value != valid_value:
            response_dict[index] = temp_dict

    response.append(response_dict)
    return response


def true_dtype(df):
    """
    return map of col unique type row wise
    """
    return {col: df[col].dropna().apply(lambda x: type(x)).unique().tolist() for col in df.columns}

def nan_check(value):
    return math.isnan(value)

def to_np_array(some_list: list):
    return np.asarray(some_list, dtype=np.float32)

def convert_kmph_to_mps(value):
    """
        convert speed in Km per hour to meter per second
    """
    return value / 3.6

def convert_time_hour_to_seconds(col):
    """
        convert pd.Series col in hour to second
    """

    return (to_np_array(col) * 3600).tolist()


def convert_distance_km_to_meter(col):
    """
        convert pd.Series col in km to meter
    """

    return (to_np_array(col) * 1000).tolist()


def set_zulu_time(now: str, offset_in_hours=None, now_format='%Y-%m-%d %H:%M:%S.%f', return_format=UTC_FORMAT):
    now = datetime.strptime(now, now_format)
    if offset_in_hours:
        now = now + timedelta(hours=offset_in_hours)
    # UTC to local time
    utc = now.replace(tzinfo=pytz.timezone('UTC'))
    local = utc.astimezone(_get_local_timezone()).replace(tzinfo=None)
    return local.strftime(return_format)


def get_utc_now():
    return str(datetime.utcnow())


def get_eta_for_utc(utc_time, time_format='%d/%m/%Y %H:%M', add_seconds=None):
    local = datetime.strptime(utc_time, UTC_FORMAT)
    if add_seconds:
        local += timedelta(seconds=add_seconds)
    return local.strftime(time_format)


def get_time_taken_for_utc(utc_veh_start_time, order_utc_time, add_seconds=None):
    veh_time = datetime.strptime(utc_veh_start_time, UTC_FORMAT)
    order_start_time = datetime.strptime(order_utc_time, UTC_FORMAT)
    if add_seconds:
        order_start_time += timedelta(seconds=add_seconds)
    calc_time = order_start_time - veh_time
    return calc_time.total_seconds()


def utc_to_local(utc_str):
    utc = datetime.strptime(utc_str, '%Y-%m-%d %H:%M:%S.%f').replace(tzinfo=pytz.timezone('UTC'))
    local = utc.astimezone(_get_local_timezone()).replace(tzinfo=None)
    return local
