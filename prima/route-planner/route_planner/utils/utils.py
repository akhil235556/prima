import itertools
import json
import logging
import math
import uuid
from datetime import datetime, timedelta
from functools import lru_cache

import numpy as np
import pandas as pd
import pytz
from dateutil import tz
from here_location_services import LS
from openpyxl.utils.cell import get_column_letter
from scipy.spatial.distance import squareform

from route_planner.constants.app_constants import HERE_API_KEY, HERE_LOCATION_TTL
from route_planner.utils.redis_cache import RedisCache
from route_planner.exceptions.exceptions import ReverseGeocodingException
from typing import Dict, Union

logger = logging.getLogger(__name__)

def generate_request_id():
    return str(uuid.uuid4()).replace('-', '')


def get_current_timestamp():
    return datetime.now(tz=tz.gettz('Asia/Kolkata'))


def dataframe_empty(df_column):
    empty_list = df_column.dropna().empty
    return empty_list


def get_now() -> datetime:
    return datetime.now(pytz.timezone('Asia/Calcutta'))


def get_eta(now: datetime, time_taken: int, time_format: str = '%d/%m/%Y %H:%M'):
    eta = now + timedelta(hours=time_taken)
    eta = eta.strftime(time_format)
    return eta

def parse_exception(rid, ex_type, value):
    ex_type = ex_type.__name__
    message = value
    error = dict(
        planning_request_id=rid,
        error_name=ex_type,
        error_message=f'{message}'
    )
    return error

def type_cast_to_str(df: pd.DataFrame, str_type_cols: list) ->None:
    if df:
        df_columns = df.columns.to_list()
        for col_name in str_type_cols:
            if col_name in df_columns:
                col = df[col_name].copy().dropna()
                col = col.apply(str)
                indexes = col.index
                df[col_name].iloc[indexes] = col


def get_column_letter_by_index(value: int) -> str:
    """
    Return excel column index for given numeric zero based index.

    Example:
        0 -> A
    """
    # Todo refactor
    # __ALPHA = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    # def hash_num_to_char(number: int) -> str:
    #     """
    #     Recursive function to get excel header by index.
    #
    #     if number is less than "26", simply hash out (index-1)
    #     Remaining possibilities,
    #     1. if remainder is zero (if quotient is 1 or not 1)
    #     2. if remainder is not zero
    #     """
    #     if number < 26:
    #         return __ALPHA[number - 1]
    #
    #     q, r = number // 26, number % 26
    #
    #     if r == 0 and q == 1:
    #         return __ALPHA[r - 1]
    #     elif r == 0:
    #         return hash_num_to_char(q - 1) + __ALPHA[r - 1]
    #     else:
    #         return hash_num_to_char(q) + __ALPHA[r - 1]
    #
    # return hash_num_to_char(value + 1)

    return get_column_letter(value + 1)


def get_header_error_response(valid_header: list, actual_header: list) -> list:
    """
    Generate Header Error Response.

    {Column index : {'Actual Header' : actual_column_name, 'Valid Header' : valid_column_name}}
    Example:
        {
            "H": {
                "Actual Header": "Unnamed: 7",
                "Valid Header": "Loading Time (Hours)"
            },
            "O": {
                "Actual Header": "Extra",
                "Valid Header": "Serviceable Vehicles"
            },
            "P": {
                "Actual Header": "Extra2",
                "Valid Header": ""
            }
        }
    """

    response = list()
    size_valid = len(valid_header)
    size_invalid = len(actual_header)
    longer_list, shorter_list = (valid_header, actual_header) if size_valid > size_invalid else (actual_header, valid_header)

    size_delta = len(longer_list) - len(shorter_list)
    shorter_list.extend([""] * size_delta)

    response_dict = dict()
    for idx, (valid_value, actual_value) in enumerate(zip(valid_header, actual_header)):
        temp_dict = dict()
        temp_dict["Actual Header"] = actual_value
        temp_dict["Valid Header"] = valid_value
        index = get_column_letter_by_index(idx)

        if actual_value != valid_value:
            response_dict[index] = temp_dict

    response.append(response_dict)
    return response


def true_dtype(df):
    """
    return map of col unique type row wise
    """
    return {col: df[col].dropna().apply(lambda x: type(x)).unique().tolist() for col in df.columns}


def nan_check(value):
    return math.isnan(value)


def to_np_array(some_list: list):
    return np.asarray(some_list, dtype=np.float32)


def unique_coordinates(coordinates):
    """
    Input to this method will be [[75.78553, 26.86891], [78.04891, 30.3097],...], list of lists of long, lats
    This method will return a list of lists with unique lat longs
    """
    coordinates.sort()
    coord = list(l for l, _ in itertools.groupby(coordinates))

    return coord

def pincode_to_latlong(locations):
    lats = []
    longs = []
    for location in locations:
        address = f"{location}, India"
        here_obj = HereAPI()
        coordinate = here_obj.get_coordinate_from_location(address)
        longs.append(coordinate.get('longitude'))
        lats.append(coordinate.get('latitude'))

    return lats, longs

def here_api_conversion(address):
    ls = LS(api_key=HERE_API_KEY)
    geo = ls.geocode(query=address)
    a = json.loads(geo.as_json_string())
    return a

def coordinates_to_matrix(coordinates):
    """
    Input to this method will be [[75.78553, 26.86891], [78.04891, 30.3097],...], list of lists of long, lats
    [75.78553, 26.86891] - 0 index is long and 1 index is lat.
    """
    place_coordinates_comb = itertools.combinations(coordinates, 2)
    place_coordinates_comb_flatten = [(*x, *y) for (x, y) in place_coordinates_comb]
    lon1, lat1, lon2, lat2 = np.array(place_coordinates_comb_flatten).T
    d_vect = haversine_np(lon1, lat1, lon2, lat2)
    place_correlation = pd.DataFrame(squareform(d_vect))
    matrix = place_correlation.values.tolist()
    for k in range(len(matrix)):
        matrix[k][0] = 0

    return matrix

def haversine_np(lon1, lat1, lon2, lat2):
    """
    Calculate the great circle distance between two points
    on the earth (specified in decimal degrees)

    All args must be of equal length.
    """
    lon1, lat1, lon2, lat2 = map(np.radians, [lon1, lat1, lon2, lat2])

    dlon = lon2 - lon1
    dlat = lat2 - lat1

    a = np.sin(dlat / 2.0) ** 2 + np.cos(lat1) * np.cos(lat2) * np.sin(dlon / 2.0) ** 2

    c = 2 * np.arcsin(np.sqrt(a))
    km = 6367 * c * 1000
    return km


def remove_empty_keys(dict):
    _dict = {}
    for k, v in dict.items():
        if v:
            _dict[k] = v
    dict = _dict
    return dict


def remove_suffix(string, suffix):
    if string.endswith(suffix):
        return string[:-len(suffix)]
    return string
class Cache:
    cache_obj = None

    def __init__(self, cache_obj=None):
        self.cache_obj = cache_obj
        self.data = dict()

    def get_cache_data(self, key):
        data = None
        if isinstance(self.cache_obj, RedisCache):
            data = self.cache_obj.get_data(key)
        return data

    def set_cache_data(self, key, expiry_time, data):
        try:
            self.cache_obj.set_key_data(data=data, key=key)
            self.cache_obj.set_expire_key(key, expiry_time)
        except Exception as e:
            logger.error(f"Caching Error {e}")


def get_here_redis_cache_key(address: str) -> str:
    """
    generate a HERE location redis cache key
    formatted_address is :
        -> strip
        -> lowercase
        -> ' ' replaced with '_'

    format: 'HERE:formatted_address'

    address: str

    """

    key = "+".join(address.strip().split(' '))
    prefix = "HERE:"
    return f"{prefix}{key.lower()}"

class HereAPI:
    _cache: Cache = None

    def __init__(self):
        redis_cache = RedisCache(data={})
        self._cache = Cache(cache_obj=redis_cache)

    def get_redis_cache_key(self, address: str) -> str:
        return get_here_redis_cache_key(address)

    def get_coordinate_from_location(self, address: str) -> Dict[str, Union[float, None]]:
        cache_key = self.get_redis_cache_key(address)

        # check Cache
        coordinate_dict = self._cache.get_cache_data(cache_key)
        if coordinate_dict is not None:
            logger.info(f"HereAPI : found {address} in cache : {coordinate_dict}")
        else:
            # Call HERE API
            logger.info(f"HereAPI : calling HERE API : {address}")
            api_response = self.here_api_conversion(address)
            coordinate_dict = self.parse_here_api_response(api_response)
            if coordinate_dict is None:
                logger.error(f"HereAPI: Response : {api_response}")
                raise ReverseGeocodingException(f"Unable to find coordinate for location: {address}")

            # Set cache
            expiry_time = HERE_LOCATION_TTL
            self._cache.set_cache_data(key=cache_key, data=coordinate_dict, expiry_time=expiry_time)
            logger.info(f"Cache Updated :: [{cache_key}] : {self._cache.get_cache_data(key=cache_key)}, TTL : {expiry_time}s")

        return coordinate_dict

    @staticmethod
    def here_api_conversion(address: str) -> dict:
        ls = LS(api_key=HERE_API_KEY)
        geo = ls.geocode(query=address)
        a = json.loads(geo.as_json_string())
        return a

    @staticmethod
    def parse_here_api_response(api_response: dict) -> Dict[str, Union[float, None]]:
        items = api_response.get('items', None)

        if not items or 'position' not in items[0] or not {'lat', 'lng'}.issubset(set(items[0]['position'].keys())):
            return None

        return {
            'latitude': items[0]['position']['lat'],
            'longitude': items[0]['position']['lng']
        }
