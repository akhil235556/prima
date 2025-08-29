from openpyxl.utils.cell import get_column_letter


def get_column_letter_by_index(value: int) -> str:
    """
    Return excel column index for given numeric zero based index.

    Example:
        0 -> A
    """
    # Todo refactor
    # __ALPHA = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    # def hash_num_to_char(number: int) -> str:
    #     """
    #     Recursive function to get excel header by index.
    #
    #     if number is less than "26", simply hash out (index-1)
    #     Remaining possibilities,
    #     1. if remainder is zero (if quotient is 1 or not 1)
    #     2. if remainder is not zero
    #     """
    #     if number < 26:
    #         return __ALPHA[number - 1]
    #
    #     q, r = number // 26, number % 26
    #
    #     if r == 0 and q == 1:
    #         return __ALPHA[r - 1]
    #     elif r == 0:
    #         return hash_num_to_char(q - 1) + __ALPHA[r - 1]
    #     else:
    #         return hash_num_to_char(q) + __ALPHA[r - 1]
    #
    # return hash_num_to_char(value + 1)

    return get_column_letter(value + 1)


def get_header_error_response(valid_header: list, actual_header: list) -> list:
    """
    Generate Header Error Response.

    {Column index : {'Actual Header' : actual_column_name, 'Valid Header' : valid_column_name}}
    Example:
        {
            "H": {
                "Actual Header": "Unnamed: 7",
                "Valid Header": "Loading Time (Hours)"
            },
            "O": {
                "Actual Header": "Extra",
                "Valid Header": "Serviceable Vehicles"
            },
            "P": {
                "Actual Header": "Extra2",
                "Valid Header": ""
            }
        }
    """

    response = list()
    size_valid = len(valid_header)
    size_invalid = len(actual_header)
    longer_list, shorter_list = (valid_header, actual_header) if size_valid > size_invalid else (actual_header, valid_header)

    size_delta = len(longer_list) - len(shorter_list)
    shorter_list.extend([""] * size_delta)

    response_dict = dict()
    for idx, (valid_value, actual_value) in enumerate(zip(valid_header, actual_header)):
        temp_dict = dict()
        temp_dict["Actual Header"] = actual_value
        temp_dict["Valid Header"] = valid_value
        index = get_column_letter_by_index(idx)

        if actual_value != valid_value:
            response_dict[index] = temp_dict

    response.append(response_dict)
    return response


def true_dtype(df):
    """
    return map of col unique type row wise
    """
    return {col: df[col].dropna().apply(lambda x: type(x)).unique().tolist() for col in df.columns}

